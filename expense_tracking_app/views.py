from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.models import Group,User,Permission
from django.contrib.auth.forms import PasswordChangeForm

from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login,logout
from django.shortcuts import redirect, render
from openpyxl import load_workbook
from django.db.models import Sum,Avg
from datetime import datetime


from .models import *
from .form import *
from .filter import BookFilter

@login_required(login_url='login')
def home(request):
    types = Book_type.objects.all()

    catagories = []
    datapoints = []
    all_books = Books.objects.all()
    total_expense= all_books.aggregate(sum = Sum('distribution_expense'))['sum']
    for type in types:
        
        books = Books.objects.filter(category = type)
        expense = books.aggregate(Sum("distribution_expense"))["distribution_expense__sum"]
        average_expense = books.aggregate(average_expense = Avg("distribution_expense"))['average_expense']
        
        category = {"name":type,"books":len(books),"expense":expense,"average" : average_expense}
        catagories.append(category)
        y = round((expense/total_expense)*100) if len(books)>0 else 0
        data = {"label":type.type_name,"y":y}
        datapoints.append(data)
        
    
    
    total = {"books":len(all_books),"category":len(catagories),"expense":total_expense,"average":all_books.aggregate(avg = Avg('distribution_expense'))['avg']}
        
    context ={"datapoints":datapoints,"catagories":catagories,"total":total}
    
    return render(request,"home.html", context)

# book
@login_required(login_url='login')
def book_management(request):
    
    books_list = Books.objects.all()  
    bookFilter = BookFilter(request.GET, queryset=books_list)
    books_list = bookFilter.qs
    book_per_page = 15
    paginator = Paginator(books_list,book_per_page)
    page_num = request.GET.get('page',1)
    try:
        books = paginator.page(page_num)
        page_offset = (int(page_num)-1)*book_per_page 
    except PageNotAnInteger:
        # if page is not an integer, deliver the first page
        books = paginator.page(1)
        page_offset = 0
        
    except EmptyPage:
        # if the page is out of range, deliver the last page
        books = paginator.page(paginator.num_pages)
        page_offset = ((paginator.num_pages)-1)*book_per_page

    
    return render(request,'book_management.html',{"books":books,"bookFilter":bookFilter,"page_offset":page_offset})
@login_required(login_url='login')
def add_book(request):
    form = Books_form()
    message = ""
    # add individually 
    if request.method == 'POST' and 'idNumber' in request.POST:
        form = Books_form(request.POST) 

        if form and  form.is_valid():
            book =  form.save( commit=False)
            book.modification_log = "today |yosef |add"
            book.save()
            message = "Done,Record saved "
            form = Books_form()
        else:
            form = Books_form(request.POST)
            message = "unable to save please try again !"



    # import from excel sheet
    if request.method == 'POST' and 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            nonImported_rows = []
        # try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    id,title	,subtitle	,authors	,publisher,	published_date,	category,	distribution_expense = row
                    category =  "- - -" if category is None  else category
                    category_typ = Book_type.objects.filter(type_name__exact = category).first()
                    if category_typ is None:
                        modification_log = datetime.now().strftime("%Y-%m-%d %H:%M:%S") +" | "+str(request.user)+" | Imported"
                        Book_type.objects.create(type_name = category,modification_log = modification_log)
                    modification_log = datetime.now().strftime("%Y-%m-%d %H:%M:%S") +" | "+str(request.user)+" | Imported"
                    Books.objects.create( idNumber =id,title =title	, subtitle =subtitle	,author =authors	, publisher =publisher,	published_date =published_date,	category=category_typ,	 distribution_expense =distribution_expense ,modification_log = modification_log)
                except ValueError as e :
                    context = {
                        "occurAt":"Importing File",
                        "header":"Invalid data Format",
                        "description":"Error in  Number of columns ," +e.args[0]+". Fix the issue on your file and try again "
                    }

                    return render(request, 'error.html',context)

                except Exception as e: 
                  #track rows that that cant be imported with their error 
                   nonImported_row = [id,title	,subtitle	,authors	,publisher,	published_date,	category,	distribution_expense,str(e)]
                   nonImported_rows.append(nonImported_row)
                   
                
            context = {"header":"Import Completed ",
                        "message":"The below rows cant be imported due to the mentioned cases",
                        'data':nonImported_rows
                        }
            
            return render(request, 'success.html',context)

    context = {
        "title":"Add Book",
        "header":"Add Book | Adding a Record ",
        "description":"Add a book, fill out all the required field, and save. You can add multiple records as you wish, and you have the option to import records from an Excel sheet by clicking the import button.",
        "submitBtn":"Save,Add Another",
        "form":form,
        "message":message,
        "add_category":add_category(request),
      
    }
    return render(request,'book_crud.html',context)
@login_required(login_url='login')
def editBook(request,id):
    book = Books.objects.get(pk = id)
    message = ""
    form = Books_form( instance=book)
    if request.method == 'POST':
        form = Books_form(request.POST, instance=book)
        if form.is_valid():
            book =  form.save( commit=False)
            book.modification_log ="negem ==="+ book.modification_log
            book.save()
            message = "Done,Record saved "
                    
            return redirect('books')  # Redirect to a success URL

        else:
            form = Books_form(instance=book)
            message = "unable to save please try again !"
    context = {
        "title":"Update Records",
        "header":"Update Records | Editing Books Data ",
        "description":"Edit book data effortlessly: simply modify the field data, save your changes, and watch as they're promptly updated in the database",
        "submitBtn":"Update",
        "form":form,
        "message":message,
         "add_category":add_category(request),
      
    }
    return render(request,'book_crud.html',context)
@login_required(login_url='login')
def deleteBook(request,id):
    book = Books.objects.get(pk = id)
    form = Books_form(instance=book)
    message = ""
    if request.method == 'POST':
        form = Books_form(request.POST, instance=book)

        try:
            book.delete()
            message = "Done,Record Deleted "
            return render(request,'success.html',{"message":message,"header":"Completed",'backto':'books'})
        except Exception as e:         
            message = "unable to save please try again !"+e

    context = {
        "title":"Delete Records",
        "header":"Delete | Erasing Books Data ",
        "description":"Erasing Books Data,Are you sure you want to delete this book? Deleting it is permanent, with no option for recovery once it's lost. If you're uncertain, please cancel the action",
        "submitBtn":"Delete ",
        "form":form,
        "message":message
      
    }
    return render(request,'book_crud.html',context)

# book category
@login_required(login_url='login')
def add_category(request):
    category_form = Category_form()
    if request.method == 'POST':
        category_form = Category_form(request.POST)
        if category_form.is_valid() :
            category = category_form.save(commit=False)
            category.modification_log = datetime.now().strftime("%Y-%m-%d %H:%M:%S") +" | "+str(request.user)+" | Create"
            category.save()
            category_form = Category_form()

    return category_form
# user
@login_required( login_url='/login')
def user_managment(request):
    users = User.objects.all()

    return render(request,'user_managment.html',{'users':users })
@login_required(login_url='login')
def add_user(request): 
    message = ""     
    
    user_form = User_form()
    profile_form = Profile_form()
    if len(Group.objects.all())<2:define_group()

    
    if request.method == 'POST':
        user_form = User_form(request.POST)
        profile_form = Profile_form(request.POST,request.FILES)
        if user_form.is_valid() and profile_form.is_valid():
            user =  user_form.save()
            user_profile = profile_form.save(commit=False)
            user.groups.add(user_profile.user_group)
            user_profile.user = user
            
            user_profile.save()
            
            user_form = User_form()
            profile_form = Profile_form()
            message = {'msg':"Done,User Created",'type':'success'}
        else:
            message = {'msg': "Unable to create user",'type':'danger'}

           
           
           
    context = {
        "title":"Add User",
        "header":"Add record | Creating New User ",
        "description":"Add a User, fill out all the required field, and save. You can add multiple records as you wish,",
        "submitBtn":"Save ",
        'user_form':user_form,
        'profile_form':profile_form,
        "message":message,
      
    }
            
    return render(request,'user_crud.html',context)
@login_required(login_url='login')
def editUser(request,id):
    user = User.objects.get(id = id)
    message = "" 
    username =""
    if request.method == 'POST':
        profile_form = Profile_form(request.POST,request.FILES,instance=user.profile)
        if profile_form.is_valid():
            if  'edit_toggle' in  request.POST:
                togglevalue  = request.POST.get('edit_toggle')
                if togglevalue == 'edited':
                    username = request.POST['username']
                    # if user.username  == username:
                    post_data = request.POST.copy()
                    username = post_data.get('username', '')
                    post_data['username'] = username + "xx"   
                    user_form = User_form(request.POST,post_data)
                    if user_form.is_valid() :
                        user.set_password( user_form.cleaned_data['password1'])
                        user.save()
                        user_profile = profile_form.save(commit=False)
                        user_profile.user = user
                        user_profile.save()
                        return redirect('users')
                    else:
                        message = {'msg': "Unable to update user",'type':'danger'}
            else:
                profile_form.save()
                return redirect('users')

        else:
            message = {'msg': "Unable to create user",'type':'danger'}
    else:
        profile_form = Profile_form( instance=user.profile)

        user_form = User_form( instance=user)
           
           
           
    context = {
        "title":"Edit User",
        "header":"Edit record |Update User ",
        "description":"Edit User data effortlessly: simply modify the field data, save your changes, and watch as they're promptly updated in the database",
        "submitBtn":"Update ",
        'user_form':user_form,
        'profile_form':profile_form,
        "message":message,
      
    }
    return render(request,'user_crud.html',context)   
@login_required(login_url='login')
def deleteUser(request,id):
    message = "" 
    user = User.objects.get(id = id)
    if request.method == 'POST':
        profile_form = Profile_form(request.POST,request.FILES,instance = user.profile)
        try:
            User.objects.get(id = id).delete()            
            message = "Done,User Account Removed"
            return render(request,'success.html',{"message":message,"header":"Completed",'backto':'users'})
            
        except Exception as a:
            message = {'msg': a,'type':'danger'}
            
    else:
        profile_form = Profile_form(instance=user.profile)
                      
    context = {
        "title":"Remove User",
        "header":"Data removing | Removing User Account ",
        "description":"Erasing User Data,Are you sure you want to delete this User? Deleting it is permanent, with no option for recovery once it's lost. If you're uncertain, please cancel the action",
        "submitBtn":"Delete ",
       
        'profile_form':profile_form,
        "message":message,
      
    }
            
    return render(request,'user_crud.html',context)

# authentication and authorization  
def user_login(request):
    error,succuss = "",""
    if request.user.is_authenticated:
        return redirect('home')

    elif request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request,username = username, password = password)
        if user is not None :  
            login(request,user)
            return redirect('home')
        else:
            error  = "username or password incorrect"
            return render(request,'login.html' ,{"error":error})
    else:
        return render(request,'login.html')
def user_logout(request):
    logout(request)
    return redirect('login')
    
def define_group():
    group_user,created =Group.objects.get_or_create(name = 'User') 
    group_admin,created =Group.objects.get_or_create(name = 'Admin') 
    # content type
    ct_book = ContentType.objects.get_for_model(Books)
    ct_book_type = ContentType.objects.get_for_model(Book_type)
    ct_user = ContentType.objects.get_for_model(User)
    # permissions
    permissions_book = Permission.objects.filter(content_type =ct_book )
    permissions_book_type = Permission.objects.filter(content_type =ct_book_type )
    permissions_user = Permission.objects.filter(content_type =ct_user )
    # set group permission 
    for permission in permissions_book:
        if permission.codename == 'view_book':
            group_user.permissions.add(permission)
        group_admin.permissions.add(permission)
    for permission in permissions_book_type:
        if permission.codename == 'view_book_type':
            group_user.permissions.add(permission)
        group_admin.permissions.add(permission)
    for permission in permissions_user:
        if permission.codename == 'view_user':
            group_user.permissions.add(permission)
        group_admin.permissions.add(permission)
    
@login_required(login_url='login')
def my_profile(request):
    password_form = PasswordChangeForm(request.user)

    return render( request,'profile.html' ,{"password_form":password_form})